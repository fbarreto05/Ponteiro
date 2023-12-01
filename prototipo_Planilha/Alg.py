from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

class BinaryTree:
    class Node:
        def __init__(self, nome, registro, entrada, saida):
            self.nome = nome
            self.registro = registro
            self.entrada = entrada
            self.saida = saida
            self.leftChild = None
            self.rightChild = None

    def __init__(self):
        self.root = None

    def insert(self, nome, registro, entrada, saida):
        node = self.Node(nome, registro, entrada, saida)
        if self.root is None:
            self.root = node
        else:
            current = self.root
            while current:
                if nome < current.nome:
                    if current.leftChild is None:
                        current.leftChild = node
                        break
                    current = current.leftChild
                else:
                    if current.rightChild is None:
                        current.rightChild = node
                        break
                    current = current.rightChild

    def traverse(self):
            return self.traverseRec(self.root)

    def traverseRec(self, node):
        if node:
            yield from self.traverseRec(node.leftChild)
            yield (node.nome, node.registro, node.entrada, node.saida)
            yield from self.traverseRec(node.rightChild)

myTree = BinaryTree()

wb = load_workbook('ponto.xlsx')
ws = wb['teste']
cell_range = ws['A2:D26']

for row in cell_range:
    myTree.insert(row[0].value, row[1].value, row[2].value, row[3].value)

for i in myTree.traverse():
    print('{', str(i),'}')
