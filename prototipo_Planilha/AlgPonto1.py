from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

def merge(lyst, copyBuffer, low, middle, high):
    i1 = low
    i2 = middle + 1 

    for i in range(low, high + 1):
        if i1 > middle:
            copyBuffer[i] = lyst[i2]
            i2 += 1
        elif i2 > high:
            copyBuffer[i] = lyst[i1]
            i1 += 1
        elif lyst[i1] < lyst[i2]:
            copyBuffer[i] = lyst[i1]
            i1 += 1
        else: 
            copyBuffer[i] = lyst[i2]
            i2 += 1
    for i in range(low, high + 1): 
        lyst[i] = copyBuffer[i]

def mergeSortHelper(lyst, copyBuffer, low, high):
    if low < high:
        middle = (low + high)//2
        mergeSortHelper(lyst, copyBuffer, low, middle)
        mergeSortHelper(lyst, copyBuffer, middle + 1, high)
        merge(lyst, copyBuffer, low, middle, high)

def mergeSort(lyst):
    copyBuffer = [0]*(len(lyst))
    mergeSortHelper(lyst, copyBuffer, 0, len(lyst)-1)

def menu():
    escolha = -1
    while (escolha < 0 or escolha > 3):
        print ("\tDIGITE A ESCOLHA\n"
            "1 - Ordem alfabética\n"
            "2 - Ordem por registro\n"
            "0 - Sair\n") 
        escolha = int(input("Digite: "))
        if (escolha > 0 and escolha > 2):
            print("Opção inválida! Digite Novamente\n") 
        return escolha

wb = load_workbook('ponto.xlsx')
ws = wb['teste']

def opcao(escolha):
    match escolha:
        case 1: 
            
            cell_range = ws['A2:D26']

            dados = [(a.value, b.value, c.value, d.value) for a, b, c, d in cell_range]
            
            mergeSort(dados)

            for i, (nome, registro, entrada, saida) in enumerate(dados):
                cell = ws.cell(row = i + 2, column=1)
                cell.value = nome
                cell = ws.cell(row = i + 2, column=2)
                cell.value = registro
                cell = ws.cell(row = i + 2, column=3)
                cell.value = entrada
                cell = ws.cell(row = i + 2, column=4)
                cell.value = saida
                wb.save('pontoOrdemAlf.xlsx')
        case 2:

            cell_range = ws['A2:D26']

            dados2 = [(b.value, a.value, c.value, d.value) for a, b, c, d in cell_range]

            mergeSort(dados2)

            for i, (registro, nome, entrada, saida) in enumerate(dados2):
                cell = ws.cell(row = i + 2, column=2)
                cell.value = registro
                cell = ws.cell(row = i + 2, column=1)
                cell.value = nome
                cell = ws.cell(row = i + 2, column=3)
                cell.value = entrada
                cell = ws.cell(row = i + 2, column=4)
                cell.value = saida
            wb.save('pontoRegistro.xlsx')
        case 3:
            exit

escolha = menu()
opcao(escolha)

