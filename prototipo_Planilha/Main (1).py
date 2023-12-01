from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from BinaryTree import BinaryTree

myTree = BinaryTree()

wb = load_workbook('ponto.xlsx')
ws = wb['teste']
cell_range = ws['A2:D26']

def menu():
    choice = 100
    while (choice < 0 or choice > 3):
        print ("\tDIGITE A ESCOLHA\n"
            "1 - Ordem alfabética\n"
            "2 - Ordem por registro\n"
            "0 - Sair\n") 
        choice = int(input("Digite: "))
        if (choice < 0 or choice > 2):
            print("Opção inválida! Digite Novamente\n") 
        else:
            return choice

def choice(choice):
    if choice == 1:
        for row in cell_range:
            myTree.insert_Alf(row[0].value, row[1].value, row[2].value, row[3].value)
    elif choice == 2:
        for row in cell_range:
            myTree.insert_Reg(row[0].value, row[1].value, row[2].value, row[3].value)

    i = 0
    for data in myTree.trasverse():
        cell = ws.cell (row = i + 2, column = 1)
        cell.value = data[0]
        cell = ws.cell (row = i + 2, column = 2)
        cell.value = data[1]
        cell = ws.cell (row = i + 2, column = 3)
        cell.value = data[2]
        cell = ws.cell (row = i + 2, column = 4)
        cell.value = data[3]
        i += 1
    if choice == 1:
        wb.save('pontoOrdem_alf.xlsx')
    elif choice == 2: 
        wb.save('pontoOrdem_reg.xlsx')

temp = menu()
choice(temp)